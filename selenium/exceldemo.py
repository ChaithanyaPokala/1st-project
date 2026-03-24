import openpyxl
book=openpyxl.load_workbook("C:\\Users\\chay\\OneDrive\\Documents\\Excel demo.xlsx")
sheet=book.active #to activate the sheet of excel
Dict={}
cell=sheet.cell(row=1,column=1)
print(cell.value)
sheet.cell(row=2,column=5).value=2005
print(sheet.cell(row=2,column=5).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)
# for i in range(1,sheet.max_row+1):#for printing all the rows and columns
#     for j in range(1,sheet.max_column+1):
#         print(sheet.cell(row=i,column=j).value)
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "chaithanya":#for printing specific row
        for j in range(1,sheet.max_column+1):
            # print(sheet.cell(row=i,column=j).value)
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value  #to store in dictionary
print(Dict)c

